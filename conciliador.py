"""
Conciliador Banco x Corporativo  —  v2
----------------------------------------
Motor de match: nome (70%) + valor com tolerância (20%) + proximidade de data (10%)
Tolerância de valor: max(R$1,00 ; 2% do valor) — captura divergências de centavos/juros
Aliases: mapeamentos conhecidos entre nomes do corporativo e do extrato

Como usar:
  1. Edite o config.yaml com os arquivos e o mês corretos.
  2. Coloque os dois arquivos Excel na mesma pasta.
  3. Abra o terminal nessa pasta e rode:  python conciliador.py
  4. Abra o arquivo de saída gerado.
"""

import pandas as pd
import re
import unicodedata
import sys
import warnings
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── config ────────────────────────────────────────────────────────────────────
def carregar_config():
    defaults = {
        "arquivo_corporativo": "Base_Corporativo.xlsx",
        "aba_corporativo":     "ControlePagamentos",
        "arquivo_banco":       "Base_BTG_-_Londrina.xlsx",
        "aba_banco":           "Extrato",
        "unidade":             "Londrina",
        "banco":               "BTG",
        "mes":                 5,
        "saida":               "Conciliacao_Londrina_Maio.xlsx",
        # parâmetros do motor de match
        "limiar_nome":         45,    # score mínimo de nome para considerar par
        "tolerancia_abs":      1.00,  # R$ de tolerância no valor
        "tolerancia_perc":     0.02,  # % de tolerância no valor (usa o maior)
        "janela_data_dias":    10,    # dias máx para pontuar proximidade de data
        # pesos do score composto (devem somar 1.0)
        "peso_nome":           0.70,
        "peso_valor":          0.20,
        "peso_data":           0.10,
    }
    try:
        import yaml
        with open("config.yaml", encoding="utf-8") as f:
            dados = yaml.safe_load(f)
        defaults.update({k: v for k, v in dados.items() if v is not None})
    except ImportError:
        print("Aviso: PyYAML não instalado — usando config interno.")
    except FileNotFoundError:
        print("Aviso: config.yaml não encontrado — usando config interno.")
    return defaults


# ── aliases: nomes do corporativo → nomes equivalentes no extrato ─────────────
# Cada entrada: (fragmento_no_nome_normalizado, substituição_normalizada)
# O fragmento é buscado no nome já normalizado (minúsculas, sem acento, sem sufixos)
ALIASES = [
    ("das simples nacional",   "ministerio da fazenda"),
    ("simples nacional",       "ministerio da fazenda"),
    ("copel distribuicao",     "copel distribuicao"),
    ("copel",                  "copel distribuicao"),
    ("claro s a",              "claro"),
    ("pagseguro internet",     "pagseguro"),
    ("stone pagamentos",       "stone"),
    ("gf marketing digital",   "gf marketing digital"),
    ("ouro verde coleta",      "ouro verde coleta"),
]


# ── normalização ─────────────────────────────────────────────────────────────
def sa(s: str) -> str:
    """Remove acentos."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(s))
        if not unicodedata.combining(c)
    )


def nn(s: str, aliases: list = ALIASES) -> str:
    """Normaliza nome para comparação e aplica aliases."""
    s = sa(s).lower()
    s = re.sub(r"\-\s*qrcode.*", "", s)           # remove sufixo QRCode
    s = re.sub(r"\b[a-z0-9]{12,}\b", "", s)       # remove hashes longos
    s = re.sub(r"\(\d+/\d+\)", "", s)              # remove "(2/4)"
    s = re.sub(r"\bltda\b|\bs\.?a\.?\b|\bepp\b|\bme\b|\bmei\b", "", s)
    s = re.sub(r"[^a-z ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for fragmento, substituto in aliases:
        if fragmento in s:
            return substituto
    return s


def v2(x) -> float | None:
    try:
        return round(abs(float(x)), 2)
    except Exception:
        return None


def inferir_metodo(desc: str) -> str:
    if sa(str(desc)).lower().startswith("pagamento de boleto"):
        return "Boleto"
    return "PIX"


# ── leitura das bases ─────────────────────────────────────────────────────────
def carregar_corporativo(cfg: dict) -> pd.DataFrame:
    try:
        c = pd.read_excel(cfg["arquivo_corporativo"], sheet_name=cfg["aba_corporativo"], header=0)
    except FileNotFoundError:
        sys.exit(f"ERRO: '{cfg['arquivo_corporativo']}' não encontrado.")

    cl = list(c.columns)
    c = c.rename(columns={
        cl[1]:  "data_lanc",
        cl[6]:  "rec",
        cl[7]:  "serv",
        cl[8]:  "forma",
        cl[9]:  "filial",
        cl[12]: "valor",
        cl[15]: "mes",
        cl[11]: "vencimento",
    })

    c["fn"] = c["filial"].apply(lambda x: sa(x).lower().strip() if pd.notna(x) else "")
    c["pn"] = c["forma"].apply(lambda x: sa(x).lower().strip() if pd.notna(x) else "")

    unidade_n = sa(cfg["unidade"]).lower().strip()
    banco_n   = sa(cfg["banco"]).lower().strip()

    sub = c[(c["fn"] == unidade_n) & (c["pn"] == banco_n) & (c["mes"] == cfg["mes"])].copy()
    if sub.empty:
        sys.exit(
            f"ERRO: nenhum registro encontrado para "
            f"{cfg['unidade']} / {cfg['banco']} / Mês {cfg['mes']}."
        )

    sub["va"]   = sub["valor"].apply(v2)
    sub["rn"]   = sub["rec"].apply(nn)
    sub["venc"] = pd.to_datetime(sub["vencimento"], errors="coerce")
    return sub.reset_index(drop=True)


def carregar_banco(cfg: dict) -> pd.DataFrame:
    try:
        raw = pd.read_excel(cfg["arquivo_banco"], sheet_name=cfg["aba_banco"], header=None)
    except FileNotFoundError:
        sys.exit(f"ERRO: '{cfg['arquivo_banco']}' não encontrado.")

    h = next(
        (i for i in range(len(raw))
         if raw.iloc[i].astype(str).str.contains("Data de lançamento", na=False).any()),
        None,
    )
    if h is None:
        sys.exit("ERRO: coluna 'Data de lançamento' não encontrada no extrato.")

    b = pd.read_excel(cfg["arquivo_banco"], sheet_name=cfg["aba_banco"], header=h)
    b = b.loc[:, ~b.columns.astype(str).str.startswith("Unnamed")].dropna(how="all")
    b.columns = ["data", "desc", "valor", "saldo"][: len(b.columns)]
    b["valor"] = pd.to_numeric(b["valor"], errors="coerce")
    b["data"]  = pd.to_datetime(b["data"], errors="coerce", dayfirst=True)

    s = b[b["valor"] < 0].copy()

    pref = [
        "pagamento de boleto enviado para ",
        "devolucao do pix enviado para ",
        "pix enviado para ",
        "ted enviado para ",
        "pagamento de conta / tributo - ",
    ]

    def extrair_rec(d: str) -> str:
        d2 = sa(str(d)).lower()
        for p in pref:
            if d2.startswith(p):
                return str(d)[len(p):].strip()
        return str(d)

    s["rb"]      = s["desc"].apply(extrair_rec)
    s["rn"]      = s["rb"].apply(nn)
    s["va"]      = s["valor"].apply(v2)
    s["interna"] = s["desc"].apply(
        lambda d: sa(cfg["unidade"]).lower() in sa(str(d)).lower()
        and "clinica capilar" in sa(str(d)).lower()
    )
    s["estorno"] = s["desc"].apply(
        lambda d: sa(str(d)).lower().startswith("devolucao")
    )
    return s.reset_index(drop=True)


# ── motor de match ────────────────────────────────────────────────────────────
def calcular_score(cr: pd.Series, br: pd.Series, cfg: dict) -> tuple | None:
    """
    Retorna (score_total, sim_nome, score_valor, score_data, diff_abs) ou None
    se o par não passa nos filtros mínimos.

    Ponderação:
        score = sim_nome * peso_nome
              + score_valor * 100 * peso_valor
              + score_data  * 100 * peso_data
    """
    # — nome —
    sim_nome = fuzz.token_set_ratio(cr["rn"], br["rn"])
    if sim_nome < cfg["limiar_nome"]:
        return None

    # — valor com tolerância —
    va_c, va_b = cr["va"], br["va"]
    if not va_c or not va_b:
        return None
    diff_abs = abs(va_c - va_b)
    tol = max(cfg["tolerancia_abs"], va_c * cfg["tolerancia_perc"])
    if diff_abs > tol:
        return None
    score_valor = max(0.0, 1.0 - diff_abs / tol)  # 1.0 = exato; decai até 0 no limite

    # — data (vencimento corp vs data banco) —
    score_data = 0.0
    if pd.notna(cr.get("venc")) and pd.notna(br.get("data")):
        dias = abs((cr["venc"] - br["data"]).days)
        score_data = max(0.0, 1.0 - dias / cfg["janela_data_dias"])

    score = (
        sim_nome      * cfg["peso_nome"]
        + score_valor * 100 * cfg["peso_valor"]
        + score_data  * 100 * cfg["peso_data"]
    )
    return round(score, 2), sim_nome, round(score_valor, 3), round(score_data, 3), diff_abs


def conciliar(corp: pd.DataFrame, banco: pd.DataFrame, cfg: dict) -> tuple:
    """
    Retorna (conciliados, divergentes, corp_sem_par, banco_sem_par).

    conciliados  : lista de (i, j, score, sim_nome, diff_abs) — valor exato
    divergentes  : lista de (i, j, score, sim_nome, diff_abs) — valor dentro da tolerância mas ≠
    corp_sem_par : índices do corp sem match
    banco_sem_par: índices do banco sem match
    """
    candidatos = []
    for i, cr in corp.iterrows():
        for j, br in banco.iterrows():
            if br["interna"] or br["estorno"]:
                continue
            res = calcular_score(cr, br, cfg)
            if res is None:
                continue
            score, sim_nome, sv, sd, diff_abs = res
            candidatos.append((score, sim_nome, sv, sd, diff_abs, i, j))

    # match 1:1 guloso (maior score primeiro)
    candidatos.sort(reverse=True)
    uc, ub, M = {}, {}, []
    for score, sim_nome, sv, sd, diff_abs, i, j in candidatos:
        if i in uc or j in ub:
            continue
        uc[i] = ub[j] = True
        M.append((score, sim_nome, sv, sd, diff_abs, i, j))

    conciliados  = [(i, j, sc, sn, da) for sc, sn, sv, sd, da, i, j in M if da == 0]
    divergentes  = [(i, j, sc, sn, da) for sc, sn, sv, sd, da, i, j in M if da > 0]
    corp_sem_par = [i for i in corp.index if i not in uc]
    banco_sem_par= [j for j in banco.index if j not in ub
                    and not banco.loc[j, "interna"] and not banco.loc[j, "estorno"]]

    return conciliados, divergentes, corp_sem_par, banco_sem_par


# ── geração do relatório ──────────────────────────────────────────────────────
AZUL     = "1F4E79"
VERM     = "C00000"
VERDE_C  = "375623"
LARANJA  = "C55A11"
AMARELO  = "FFF2CC"

def _cabecalho(ws, ncol: int, titulo: str, cor: str = AZUL):
    ws.append([titulo])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncol)
    c = ws.cell(ws.max_row, 1)
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ws.max_row].height = 22


def _linha_col(ws, cols: list):
    ws.append(cols)
    for k in range(1, len(cols) + 1):
        c = ws.cell(ws.max_row, k)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _aba(wb, nome: str, titulo: str, cabec: list, linhas: list, cor: str = AZUL):
    w = wb.create_sheet(nome)
    _cabecalho(w, len(cabec), titulo, cor)
    w.append([])
    _linha_col(w, cabec)
    for ln in linhas:
        w.append(ln)
        r = w.max_row
        for k in range(1, len(cabec) + 1):
            w.cell(r, k).font      = Font(name="Arial")
            w.cell(r, k).alignment = Alignment(vertical="center", wrap_text=True)
    for k in range(1, len(cabec) + 1):
        max_w = max(
            [len(str(cabec[k - 1]))]
            + [len(str(w.cell(rr, k).value or "")) for rr in range(4, w.max_row + 1)]
        )
        w.column_dimensions[get_column_letter(k)].width = min(max(max_w + 2, 10), 52)
    return w


def _aba_para_lancar(wb, cfg: dict, banco: pd.DataFrame, banco_nm: list):
    """Aba no padrão do corporativo para copiar e colar."""
    COLS = [
        "", "Data lançam.", "Lançado por:", "Metodo PG",
        "Classificação Antiga", "Classificação\nPnL & FCF (Proposta RCT)",
        "Nome da empresa recebedora", "Serviço prestado",
        "Forma de pagamento", "Filial pagadora",
        "Será rateado para demais unidades?", "Data Vencimento",
        "Valor item", "Valor NF", "Status Aprovação\n(Birro)", "MÊS",
    ]
    MANUAL = {2, 4, 5, 7, 14}

    w = wb.create_sheet("Para Lançar no Corp", 1)

    w.merge_cells("A1:P1")
    c = w["A1"]
    c.value = (
        "✅  Copie as linhas abaixo e cole na aba 'ControlePagamentos' da base do corporativo.  "
        "Células em AMARELO precisam ser preenchidas antes de colar."
    )
    c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    w.row_dimensions[1].height = 24

    w.append([""] * len(COLS))
    r_hdr = w.max_row
    for k, col in enumerate(COLS, 1):
        c = w.cell(r_hdr, k)
        c.value     = col
        c.font      = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    w.row_dimensions[r_hdr].height = 30

    thin  = Side(style="thin", color="D9D9D9")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j in banco_nm:
        br = banco.loc[j]
        row_vals = [
            "", br["data"], "", inferir_metodo(br["desc"]),
            "", "", br["rb"], "", cfg["banco"], cfg["unidade"],
            "Não", br["data"], br["va"], br["va"], "", cfg["mes"],
        ]
        w.append(row_vals)
        r = w.max_row
        for k in range(1, len(COLS) + 1):
            cell = w.cell(r, k)
            cell.font      = Font(name="Arial", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = borda
            if (k - 1) in MANUAL:
                cell.fill = PatternFill("solid", fgColor=AMARELO)
            if k in (2, 12):
                cell.number_format = "DD/MM/YYYY"
            if k in (13, 14):
                cell.number_format = "#,##0.00"

    larguras = [4, 14, 14, 10, 32, 28, 30, 24, 14, 12, 10, 14, 13, 13, 12, 7]
    for k, wc in enumerate(larguras, 1):
        w.column_dimensions[get_column_letter(k)].width = wc

    if banco_nm:
        r_tot = w.max_row + 1
        w.cell(r_tot, 12).value        = f"TOTAL: {len(banco_nm)} lançamentos"
        w.cell(r_tot, 12).font         = Font(name="Arial", bold=True, color=AZUL)
        w.cell(r_tot, 13).value        = round(banco.loc[banco_nm, "va"].sum(), 2)
        w.cell(r_tot, 13).font         = Font(name="Arial", bold=True, color=AZUL)
        w.cell(r_tot, 13).number_format = "#,##0.00"

    return w


def gerar_relatorio(
    cfg: dict,
    corp: pd.DataFrame,
    banco: pd.DataFrame,
    conciliados: list,
    divergentes: list,
    corp_nm: list,
    banco_nm: list,
) -> str:

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    _cabecalho(ws, 5, f"Conciliação — {cfg['unidade']} / {cfg['banco']} / Mês {cfg['mes']}")
    ws.append([])

    # Parâmetros do motor (auditoria)
    ws.append(["Parâmetros do motor de match"])
    ws.cell(ws.max_row, 1).font = Font(name="Arial", italic=True, color="595959")
    params = (
        f"Limiar nome: {cfg['limiar_nome']}  |  "
        f"Tolerância valor: R${cfg['tolerancia_abs']:.2f} ou {cfg['tolerancia_perc']*100:.0f}%  |  "
        f"Janela data: {cfg['janela_data_dias']} dias  |  "
        f"Pesos: nome {cfg['peso_nome']:.0%} / valor {cfg['peso_valor']:.0%} / data {cfg['peso_data']:.0%}"
    )
    ws.append([params])
    ws.cell(ws.max_row, 1).font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.append([])

    _linha_col(ws, ["Categoria", "Qtd", "Valor (R$)", "Observação", "O que fazer"])

    linhas_res = [
        (
            "✓  Conciliados — valor e nome batem exatamente",
            len(conciliados),
            round(sum(corp.loc[i, "va"] for i, j, *_ in conciliados), 2),
            "",
            "Nenhuma ação necessária",
        ),
        (
            "~  Conciliados com divergência de valor — mesmo recebedor, centavos diferem",
            len(divergentes),
            round(sum(banco.loc[j, "va"] for i, j, *_ in divergentes), 2),
            f"Tolerância aplicada: R${cfg['tolerancia_abs']:.2f} / {cfg['tolerancia_perc']*100:.0f}%",
            "Confirmar qual valor é o correto",
        ),
        (
            "⚠  A REPORTAR — pago no banco, sem registro no corporativo",
            len(banco_nm),
            round(banco.loc[banco_nm, "va"].sum(), 2),
            "",
            "Enviar ao corporativo para lançamento (ver aba 'Para Lançar no Corp')",
        ),
        (
            "⚠  Pendente — registrado no corp., sem pagamento no extrato",
            len(corp_nm),
            round(corp.loc[corp_nm, "va"].sum(), 2),
            "",
            "Verificar se foi pago via cartão / outra conta / ainda não pago",
        ),
        (
            "—  Transferências internas (excluídas da conciliação)",
            int(banco["interna"].sum()),
            round(banco.loc[banco["interna"], "va"].sum(), 2),
            "",
            "Movimentação entre contas próprias — ignorada",
        ),
    ]

    cores = {2: "FFF2CC", 3: "FFF2CC"}
    for idx, (cat, q, val, obs, acao) in enumerate(linhas_res):
        ws.append([cat, q, val, obs, acao])
        r = ws.max_row
        ws.cell(r, 3).number_format = "#,##0.00"
        for k in range(1, 6):
            ws.cell(r, k).font      = Font(name="Arial")
            ws.cell(r, k).alignment = Alignment(wrap_text=True, vertical="center")
            if idx in cores:
                ws.cell(r, k).fill = PatternFill("solid", fgColor=cores[idx])

    for col, w in zip("ABCDE", [52, 8, 16, 44, 46]):
        ws.column_dimensions[col].width = w

    # ── aba "Para Lançar no Corp" ──
    _aba_para_lancar(wb, cfg, banco, banco_nm)

    # ── A Reportar ──
    linhas = [
        [
            str(banco.loc[j, "data"])[:10],
            banco.loc[j, "rb"],
            banco.loc[j, "va"],
            banco.loc[j, "desc"],
            "A investigar",
            "",
        ]
        for j in banco_nm
    ]
    w = _aba(
        wb, "A Reportar",
        "PAGOS NO BANCO SEM REGISTRO NO CORPORATIVO  →  reportar ao corporativo",
        ["Data", "Recebedor (banco)", "Valor (R$)", "Descrição original", "Status", "Observação"],
        linhas, VERM,
    )
    for r in range(4, w.max_row + 1):
        w.cell(r, 3).number_format = "#,##0.00"

    # ── Divergência de Valor ──
    linhas = [
        [
            corp.loc[i, "rec"],
            corp.loc[i, "va"],
            banco.loc[j, "va"],
            round(banco.loc[j, "va"] - corp.loc[i, "va"], 2),
            str(banco.loc[j, "data"])[:10],
            banco.loc[j, "rb"],
            int(sn),
            round(sc, 1),
            "A confirmar",
        ]
        for i, j, sc, sn, da in divergentes
    ]
    w = _aba(
        wb, "Divergencia de Valor",
        "MESMO RECEBEDOR, VALOR DIFERENTE (dentro da tolerância)  →  confirmar",
        [
            "Recebedor (corp)", "Valor corp (R$)", "Valor banco (R$)", "Diferença (R$)",
            "Data banco", "Recebedor (banco)", "Sim. nome %", "Score", "Status",
        ],
        linhas, LARANJA,
    )
    for r in range(4, w.max_row + 1):
        for k in (2, 3, 4):
            w.cell(r, k).number_format = "#,##0.00"

    # ── Pendente no Banco ──
    linhas = [
        [corp.loc[i, "rec"], corp.loc[i, "serv"], corp.loc[i, "va"], "A investigar", ""]
        for i in corp_nm
    ]
    _aba(
        wb, "Pendente no Banco",
        "REGISTRADO NO CORPORATIVO SEM PAGAMENTO NO EXTRATO",
        ["Recebedor (corp)", "Serviço", "Valor (R$)", "Status", "Observação"],
        linhas,
    )

    # ── Conciliados ──
    linhas = [
        [
            corp.loc[i, "rec"],
            corp.loc[i, "va"],
            str(banco.loc[j, "data"])[:10],
            banco.loc[j, "rb"],
            int(sn),
            round(sc, 1),
        ]
        for i, j, sc, sn, da in conciliados
    ]
    _aba(
        wb, "Conciliados",
        "CONFERIDOS — valor e recebedor batem (auditoria)",
        ["Recebedor (corp)", "Valor (R$)", "Data banco", "Recebedor (banco)", "Sim. nome %", "Score"],
        linhas, VERDE_C,
    )

    # ── Transferências Internas ──
    linhas = [
        [str(banco.loc[j, "data"])[:10], banco.loc[j, "rb"], banco.loc[j, "va"], banco.loc[j, "desc"]]
        for j in banco.index if banco.loc[j, "interna"]
    ]
    _aba(
        wb, "Transf Internas",
        "IGNORADAS — movimentação entre contas próprias",
        ["Data", "Recebedor", "Valor (R$)", "Descrição original"],
        linhas,
    )

    wb.save(cfg["saida"])
    return cfg["saida"]


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    cfg = carregar_config()

    print(f"\nConciliando: {cfg['unidade']} / {cfg['banco']} / Mês {cfg['mes']}")
    print(f"Corporativo : {cfg['arquivo_corporativo']}")
    print(f"Banco       : {cfg['arquivo_banco']}")
    print(
        f"Motor       : nome {cfg['peso_nome']:.0%} | "
        f"valor {cfg['peso_valor']:.0%} (tol ±R${cfg['tolerancia_abs']:.2f}/{cfg['tolerancia_perc']*100:.0f}%) | "
        f"data {cfg['peso_data']:.0%} (janela {cfg['janela_data_dias']}d)\n"
    )

    corp  = carregar_corporativo(cfg)
    banco = carregar_banco(cfg)
    conc, dv, corp_nm, banco_nm = conciliar(corp, banco, cfg)

    saida = gerar_relatorio(cfg, corp, banco, conc, dv, corp_nm, banco_nm)

    print("=" * 60)
    print(f"  Conciliados (exatos)         : {len(conc):>4}  R$ {sum(corp.loc[i,'va'] for i,j,*_ in conc):>12,.2f}")
    print(f"  Conciliados (div. de valor)  : {len(dv):>4}  R$ {sum(banco.loc[j,'va'] for i,j,*_ in dv):>12,.2f}")
    print(f"  A REPORTAR (banco sem corp)  : {len(banco_nm):>4}  R$ {banco.loc[banco_nm,'va'].sum():>12,.2f}")
    print(f"  Pendente   (corp sem banco)  : {len(corp_nm):>4}  R$ {corp.loc[corp_nm,'va'].sum():>12,.2f}")
    print(f"  Transf. internas excluídas   : {int(banco['interna'].sum()):>4}  R$ {banco.loc[banco['interna'],'va'].sum():>12,.2f}")
    print("=" * 60)
    print(f"\nArquivo gerado: {saida}\n")


if __name__ == "__main__":
    main()